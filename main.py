import pandas as pd
import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# Function to parse GNSS data
def parse_gnss(gnss_data):
    gnss_lock_periods = []
    current_lock_start = None
    flag = 0
    

    for line in gnss_data:
        if '$GPRMC' in line:
            parts = line.split(',')
            if len(parts) > 1 and 'A' in parts[2] and current_lock_start == None:
                # This line indicates the start of a GNSS lock
                timestamp_str = line[1:24].strip('[]')
                try:
                    timestamp = datetime.datetime.strptime(timestamp_str, '%Y-%m-%d %H:%M:%S.%f')
                    current_lock_start = timestamp
                except ValueError:
                    # Handle lines that cannot be parsed into datetime
                    continue
                flag = 1
                flag_time = timestamp
            elif ',V,' in line:
                # This line indicates the end of a GNSS lock
                timestamp_str = line[1:24].strip('[]')
                flag = 0
                try:
                    timestamp = datetime.datetime.strptime(timestamp_str, '%Y-%m-%d %H:%M:%S.%f')
                    if current_lock_start:
                        gnss_lock_periods.append((current_lock_start, timestamp))
                        current_lock_start = None
                except ValueError:
                    # Handle lines that cannot be parsed into datetime
                    continue

    if flag:
        gnss_lock_periods.append((current_lock_start, flag_time))
    return gnss_lock_periods

# Function to parse AT_GNSS data
def parse_at_gnss(at_gnss_data):
    at_commands = []
    for line in at_gnss_data:
        if 'at+qgpsloc=2' in line:
            timestamp_str = line.split(']')[0].strip('[')
            timestamp = datetime.datetime.strptime(timestamp_str, '%Y-%m-%d_%H:%M:%S:%f')
            if '+QGPSLOC:' in at_gnss_data[at_gnss_data.index(line) + 3]:
                at_commands.append((timestamp, 'Y'))
            elif '+CME ERROR:' in at_gnss_data[at_gnss_data.index(line) + 3]:
                at_commands.append((timestamp, 'X'))
    return at_commands

# Function to determine GNSS status for a given timestamp
def get_gnss_status(timestamp, gnss_lock_periods):
    for start, end in gnss_lock_periods:
        if start <= timestamp <= end:
            return 'A', f"{start} - {end}"
    return 'V', None

# Function to create Excel file with conditional formatting
def create_excel(gnss_lock_periods, at_commands):
    # Create DataFrame
    df = pd.DataFrame(at_commands, columns=['Timestamp', 'AT_Status'])
    df['GNSS_Status(NMEA A/V)'] = df['Timestamp'].apply(lambda ts: get_gnss_status(ts, gnss_lock_periods)[0])
    df['GNSS_Period'] = df['Timestamp'].apply(lambda ts: get_gnss_status(ts, gnss_lock_periods)[1])

    # Create Excel workbook and sheets
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Write column headers
    headers = ['Timestamp', 'AT_Status', 'GNSS_Status(NMEA A/V)', 'GNSS_Period']
    for col_num, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_num, value=header)

    # Fill in the data
    for r, row in enumerate(df.itertuples(), start=2):  # Start from row 2 to account for header
        for c, value in enumerate(row[1:], start=1):  # Skip the first item because it is the Index
            ws.cell(row=r, column=c, value=str(value))

    # Initialize color counts
    green_count = 0
    yellow_count = 0
    orange_count = 0
    red_count = 0

    # Apply conditional formatting
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4):
        at_status = row[1].value  # AT_Status in the second column
        gnss_status = row[2].value  # GNSS_Status in the third column
        fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')  # Default white background
        if at_status == 'Y' and gnss_status == 'A':
            fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Green
            green_count += 1
        elif at_status == 'X' and gnss_status == 'V':
            fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')  # Yellow
            yellow_count += 1
        elif at_status == 'X' and gnss_status == 'A':
            fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')  # Orange
            orange_count += 1
        elif at_status == 'Y' and gnss_status == 'V':
            fill = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')  # Red
            red_count += 1
        for cell in row:
            cell.fill = fill

    # Adjust column width for readability
    for col in ws.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[get_column_letter(column)].width = adjusted_width

    # Create summary sheet
    ws_summary = wb.create_sheet(title="Summary")
    summary_data = [
        ('GNSS good and get location', green_count, 'C6EFCE'),
        ('GNSS bad and didn\'t get location', yellow_count, 'FFFF99'),
        ('GNSS good and didn\'t get location', orange_count, 'FFA500'),
        ('GNSS bad and get location', red_count, 'FF9999')
    ]

    for row, (text, count, color) in enumerate(summary_data, start=1):
        cell_text = ws_summary.cell(row=row, column=1, value=text)
        cell_count = ws_summary.cell(row=row, column=2, value=count)
        cell_text.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        cell_count.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

    # Save workbook
    wb.save("GNSS_AT_Status.xlsx")

# Example usage with file reading
with open('GNSS.txt', 'r') as file:
    gnss_data = file.readlines()

with open('AT_GNSS.txt', 'r') as file:
    at_gnss_data = file.readlines()

gnss_lock_periods = parse_gnss(gnss_data)
at_commands = parse_at_gnss(at_gnss_data)
create_excel(gnss_lock_periods, at_commands)
print("Done")